# config
from os.path import dirname
root_dir = dirname(__file__)
from os.path import join as opj
all_scripts  = opj(root_dir, 'Scripts')
import sys
sys.path.append(root_dir)
import sys
import openpyxl
import traceback
import multiprocessing
import threading
import Weather_download
import os
import time
import glob
import Model
import weather_manager
from soilmanager import Replace_Soilprofile2
from Weather_download import daymet_bylocation
from weather_manager import Weather2
from soilmanager import Replace_Soilprofile2
import soilmanager
from Model import runAPSIM2
import numpy as np
import arcpy
import time
import pandas as pd
import logging
#from pyproj import transform, CRS, Transformer
from arcpy.sa import *
logger = logging.getLogger(__name__)
# Check out the ArcGIS Spatial Analyst extension license
arcpy.CheckOutExtension("Spatial")

#path for the apsimx file
seriesnames=['Clarion','Canisteo', 'Coland',  'Muscatine', 'Okoboji', 'Nicollet',  "Colo", 'Gara','Tama', 'Webster', 'Buckney',
'Ackmore', 'Nodaway', 'Colo']
series = {'Clarion': (-93.880227,42.049705), 'Okoboji': (-93.900227,42.049705), 'Webster':(-93.900227,42.049705),
          'Canisteo': (-93.900227,42.04970), 'Nicollet': (-93.835782,42.078536), 'Buckney':(-93.890732, 42.041544),
           'Coland': (-93.890732,42.041544), "Colo": (-93.060069, 41.67018), 'Ackmore':(-93.060069, 41.67018),
            'Downs': (-93.141639,   41.605440),'Gara': (-93.030585, 41.652741), 'Nodaway': (-93.074647,  41.606982),
            'Muscatine': (-93.044683, 41.652082), 'Tama':(-93.042241, 41.652082), 'Colo':(-92.955551,  41.606709)}
import multiprocessing

#Spool = multiprocessing.Pool(processes=multiprocessing.cpu_count() - 2, maxtasksperchild=3)
def worker(basefile, lonlat, startyear, endyear, array, index, report = 'MaizeR'):
      try:
        df = {}
        path2apsimx = Replace_Soilprofile2(basefile, 'domtcp', lonlat, crop = None)
        # download weather data from daymet returns path2file
        weatherpath = daymet_bylocation(lonlat, startyear, endyear)
        wp = Weather2(path2apsimx, weatherpath)
        editedapsimx  = wp.ReplaceWeatherData()
        dat = runAPSIM2(editedapsimx, select_report  = report)
        df['Yield']= dat.Yield.mean()
        df['N20'] = dat.TopN2O.mean()
        df['AGB'] = dat.AGB.mean()
        df['OBJECTID'] = array[index][0]
        df['Shape'] = array[index][1]
        df["CompName"] = array[index][11]
        df["gridcode"] = array[index][3]
        df["MUKEY"] = array[index][4]
        print(df)
        db = editedapsimx.split("apsimx")[0] + "db"
        del dat
        os.remove(db)
        os.remove(editedapsimx)
        os.remove(weatherpath)
        return df
      except:
        pass
#now call list comprehension
# dd = []
# lonlat = array["Shape"]
# m= 0
# start = time.perf_counter()
# for i in range(80):
#   dd.append(worker(basefile, lonlat[i], startyear, endyear, array, i, report = 'MaizeR') )
#   m +=i
#   print(m)
# end = time.perf_counter()
# print(f"time taken is: {end -start}")
#
def Merge(dict1, dict2):
    return(dict2.update(dict1))
##dd = []
##for i in seriesnames:
##  dd.append(worker(i))
#arcpy.env.scratchWorkspace = "in_memory"
#arcpy.env.workspace = r'C:\Users\rmagala\Box\ACPF_MyProject\ACPF_DATA\sievers_case_study\soilMudCreek_20221110\RMagala_20221110\soils_MudCreek_Sievers.gdb'
#arcpy.env.outputCoordinateSystem = arcpy.SpatialReference(4326)
arcpy.env.overwriteOutput = True
class SoilRasterManagement:
  def __init__(self, in_raster, horizontable, rasterfield = "MUKEY"):
    self.raster = in_raster
    self.field = rasterfield
    self.Horizontable = horizontable
    arcpy.env.workspace = os.getcwd()
  def Organise_soils(self, cellsize =30):
    try:
      start = time.perf_counter()
      shapename = r'in_memory/shapelayer'
      geodata = 'Gis_result_geodatabase.gdb'
      if not arcpy.Exists(geodata):
       arcpy.CreateFileGDB_management(os.getcwd(), geodata)
      arcpy.env.workspace  = geodata
      if arcpy.Exists(shapename):
        arcpy.management.Delete(shapename)
      shapename = r'in_memory/shapelayer'
      #arcpy.AddMessage('converting raster to polygon')
      self.rasterpolygon =  arcpy.conversion.RasterToPolygon(self.raster , shapename, simplify ="NO_SIMPLIFY", raster_field =self.field)
      
      name = r'in_memory/feature_to_point'
      if arcpy.Exists(name):
        arcpy.management.Delete(name)
      # redefine it
      name = r'in_memory/feature_to_point'
      print('converting polygon to points')
      #arcpy.AddMessage('converting Polygon to points')
      # confirm to points
      self.rasterpoints = arcpy.management.FeatureToPoint(self.rasterpolygon , name, 'INSIDE')
      inFeatures = self.rasterpoints
      joinTable = self.Horizontable
      joinField = "MUKEY"
      expression = "CompKind = 'Series'"
      outFeature = "spatialsoilhorizon"
      print("joining tables")
      self.joinedtable = arcpy.management.AddJoin(inFeatures, joinField, joinTable, 
                                                  joinField)
      print("Selecting only soil series") 
      selectionname = "Series"
      # select only points with valid componet names
      arcpy.env.scratchWorkspace = 'in_memory'
      self.soilbyseries = arcpy.Select_analysis(self.joinedtable, where_clause = 'CompKind = \'' + selectionname + '\'')
      
      # covert to numpyaray. it is 100 times faster
      sr = arcpy.SpatialReference(4326)
      print("Converting to structured numpy array")
      #arcpy.AddMessage("Converting to structured numpy array")
      # searhc the fields
      listfield = []
      lf = arcpy.ListFields(self.soilbyseries)
      for i in lf:
        listfield.append(i.name)
      # cordnates are in position number two
      
      self.feature_array  = arcpy.da.FeatureClassToNumPyArray(self.soilbyseries,  listfield[:13], spatial_reference = sr)
      print(len(self.feature_array.dtype.descr))
      return self
    except:
           tb = sys.exc_info()[2]
           tbinfo = traceback.format_tb(tb)[0]
      
           # Concatenate information together concerning the error into a message string
           pymsg = "PYTHON ERRORS:\nTraceback info:\n" + tbinfo + "\nError Info:\n" + str(sys.exc_info()[1])
           print(pymsg + "\n")
      
           if arcpy.GetMessages(2) not in pymsg:
              msgs = "ArcPy ERRORS:\n" + arcpy.GetMessages(2) + "\n"
              arcpy.AddError(msgs)
              print(msgs)
      
    finally:
            end = time.perf_counter()
            #arcpy.AddMessage(f'conversion took: {end-start} seconds')
            print(f'Conversion took: {end-start} seconds')
# ap = SoilRasterManagement('gSSURGO', 'SurfHrz070801030303') # if it from geodatabase
# pp =ap.Organise_soils()

# get results from a pool asyn class
class Result():
    def __init__(self):
        self.val = None

    def update_result(self, val):
        self.val = val

result = Result()

def f(x):
    return x*x

#pool.apply_async(f, (10,), callback=result.update_result)
#expand grid fucntion
from itertools import product
# creating grid expansion
def expand_grid(dictionary):
   """Create a dataframe from every combination of given values."""
   return pd.DataFrame([row for row in product(*dictionary.values())], 
                       columns=dictionary.keys())
def makedf(listitem):
  import pandas
  free = []
  for i in listitem:
    if i !=None:
     free.append(i)
  df = pandas.DataFrame.from_dict(free)
  return df
# this function cleans up met, apsimx and met files after simulations
def removefiles(iterable):

    for i in iterable:
      os.remove(i)
  
# get the main directory
def CLeaUp(ws):
    getcwdir  = os.getcwd()
    if os.getcwd() != ws:
      os.chdir(ws)
    # delete apsimx files
    dbfiles = glob.glob1(ws, "*.db")
    removefiles(dbfiles)
    apsimfiles = glob.glob1(ws, "*.apsimx")
    removefiles(apsimfiles)
    # delete db files
    
    #delete excel files
    excelfiles= glob.glob1(ws, "*.csv")
    removefiles(excelfiles)
    # del weather
    wther = os.path.join(ws, 'weatherdata')
    os.chdir(wther)
    metfiles = glob.glob1(wther, "*met")
    removefiles(metfiles)
    print("clean up successful")
    # reset back to the previous directory
    os.chdir(getcwdir)
# evaluate lists
def notin(targetlist, compare):
  py = []
  for i in compare:
    if i not in targetlist:
      py.append(i)
      #print(i)
  return py
def changedtype(data):
  dt = data.dtype
  dt = dt.descr # this is now a modifiable list, can't modify numpy.dtype
  # change the type of the first col:
  dt[0] = (dt[0][0], 'float64')
  dt = numpy.dtype(dt)
  # data = numpy.array(data, dtype=dt) # option 1
  data = data.astype(dt)
  
def worker(index, basefile, start, end, array):
      try:
        df = {}
        array = ar1
        report  =reported
        start= int(startyear)
        end = int(endyear)
        basefile = basefile
        path2apsimx = Replace_Soilprofile2(basefile, 'domtcp', array[index][1], filename = array[index][2], gridcode = str(array[index][0]), Objectid = str(array[index][2]), crop = None)
        # download weather data from daymet returns path2file
        weatherpath = daymet_bylocation(ar1[index][1], start, end)
  
        wp = Weather2(path2apsimx, weatherpath)
        
        editedapsimx  = wp.ReplaceWeatherData()
       
        return editedapsimx
      except:
        pass
def Mainrun(): 
       pr = [worker(i, basefile, start, end, array) for i in a]
       listp = []
       for i in pr:
        if i != None:
          listp.append(i)
       return listp
def MainrunforMP(index): # this index is gonna come from the list a
       pr = worker(index, basefile, start, end, array)
       return pr   
def CollectforMaize(apsimx):
    try:
        df = {}
        report = "MaizeR"
        dat = runAPSIM2(apsimx)
        df['longitude'] = dat["MaizeR"].longitude.values[1]
        df["Latitude"] = dat["MaizeR"].latitude[0]
        df['OBJECTID'] = dat["MaizeR"].OBJECTID.values[0]
        df["MUKEY"] = dat["MaizeR"].soiltype.values[1].split(":")[1]
        df["CompName"] = dat["MaizeR"].soiltype.values[1].split(":")[0]
        df['meanMaizeYield']= dat["MaizeR"].Yield.mean()
        df['meanMaizeAGB'] =  dat["MaizeR"].AGB.mean()
        df["ChangeINCarbon"]    =dat['Carbon'].changeincarbon[0]
        df['meanN20'] = dat["MaizeR"].TopN2O.mean()
        df["Soiltype"] = dat["MaizeR"].soiltype.values[1]
        df["CO2"] = dat['Annual'].Top_respiration.mean()
        df["meanSOC1"] = dat['Annual'].SOC1.mean()
        df["meanSOC2"] = dat['Annual'].SOC2.mean()
        
        return df
    except:
      pass
def runapsimx(apsimx): 
 try:
    df = {}
    report = "MaizeR"
    dat = runAPSIM2(apsimx)
    df['OBJECTID'] = dat["MaizeR"].OBJECTID.values[0]
    df["Shape"] = dat["MaizeR"].longitude.values[1], dat["MaizeR"].latitude[0]
    df["MUKEY"] = dat["MaizeR"].soiltype.values[1].split(":")[1]
    df["CompName"] = dat["MaizeR"].soiltype.values[1].split(":")[0]
    df['meanMaizeYield']= dat["MaizeR"].Yield.mean()
    df['meanN20'] = dat["MaizeR"].TopN2O.mean()
    df['meanMaizeAGB'] =  dat["MaizeR"].AGB.mean()
    df["Soiltype"] = dat["MaizeR"].soiltype.values[1]
    df['longitude'] = dat["MaizeR"].longitude.values[1]
    df["Latitude"] = dat["MaizeR"].latitude[0]
    df['OBJECTID'] = dat["MaizeR"].OBJECTID.values[0]
    df["ChangeINCarbon"]    =dat['Carbon'].changeincarbon[0]
    df["RyeBiomass"] = dat["WheatR"].AGB.mean()
    df["CO2"] = dat['Annual'].Top_respiration.mean()
    df["meanSOC1"] = dat['Annual'].SOC1.mean()
    df["meanSOC2"] = dat['Annual'].SOC2.mean()
    return df
 except:
  pass
def run_MultiPros(function, variables):
    """<function, variables> Execute a process on multiple processors.
    INPUTS:
    function(required) Name of the function to be executed.
    variables(required) Variable to be passed to function.
    Description: This function will run the given fuction on to multiprocesser. Total number of jobs is equal to number of variables.        
    """
    with Pool(processes=20) as pool:
        #pp = [pool.(function, (i,)).get() for i in variables] #apply_async takes on multiple arguments in a tupply form
        pp =[]
        for i in variables:
           p=  pool.apply_async(function, (i,)).get()
           pp.append(p)
        px = utils1.makedf(pp)
        print(px)
        # run again
        resultsdir = os.path.join(os.getcwd(), "SimulationResults")
        sim = 'Simulaytedresultsx.csv'
        results = os.path.join(resultsdir, sim)
        if os.path.exists(results):
            os.remove(results)
        results = os.path.join(resultsdir, sim)
        px.to_csv(results)
        return px
        s = time.perf_counter()
def DownloadMultipleweatherfiles(function, variables):
        """<function, variables> Execute a process on multiple processors.
        INPUTS:
        function(required) Name of the function to be executed.
        variables(required) Variable to be passed to function.
        Description: This function will run the given fuction on to multiprocesser. Total number of jobs is equal to number of variables.        
        """
        with Pool(processes=20) as pool:
            mpweather= pool.map(function, variables, chunksize = 50)
            listmp = []
            for i in mpweather:
                if i != None:
                  listmp.append(i)
            return listmp


def convertpoint_to_fc(path2sim, geodatabase, nameof_featureclass ="simulatedresult_feature_class"):
    """
    Path: complete path and file name to the simulated results
    geodatabase: geodatabase path for storing the new feature class
    nameof_featureclass: name for the new feature class
    
    """
    df = pd.read_csv(path2sim)
    fd  = df.convert_dtypes()
    structuredNumpy_records = df.to_records(index= False)
    name = os.path.join(geodatabase, nameof_featureclass)
    sr = arcpy.SpatialReference(4326)
    # change some unassigned data type
    dt = structuredNumpy_records.dtype.descr
    dt[8] = ('Soiltype', '<U25')
    dt[4] = ('CompName', '<U25')
    featuretobe = structuredNumpy_records.astype(dt)
    out = arcpy.da.NumPyArrayToFeatureClass(featuretobe, name, ["longitude","Latitude"], sr)
    copyfeature  = "copyfeatureclass"
    arcpy.management.CopyFeatures(name, copyfeature)

class ReadExcel:
      def __init__(self, path, excelname):
        '''
        costructs class for creating lookup info
        paramters
        ------------
        
        excelname: name with the land use and their corresponding nesting weights
        
        parameters
        '''
        fullpath = os.path.join(path, excelname)
        if (os.path.isfile(fullpath)) and (fullpath.endswith('.xlsx')):
            self.ExcelName =  excelname
            self.path  = path
            self.fullpath = fullpath
        else:
          logging.error("File does not exists make sure it is the specified in the working directory and no csv file is allowed")
      
      def ExtractInfo(self):
        '''
        
        '''
        try:
            start = time.perf_counter()
            self.objectid = []
            self.lon =[]
            self.lat  = []
              #load the excel book
            book = openpyxl.load_workbook(self.fullpath)
          # get the sheet names 
            sheet = book.sheetnames
            for index, sheets in enumerate(sheet):
             if sheets =="Lookup_info":
               lookupinfo = book.worksheets[index]
               for i in range(1, lookupinfo.max_row+1):
                    if i !=1:
                      self.objectid.append(lookupinfo.cell(row = i, column = 1).value)
                      self.lon.append(lookupinfo.cell(row =i, column = 2).value)
                      self.lat.append(lookupinfo.cell(row =i, column = 3).value)  
            if "Lookup_info" not in sheet:
              print("No look up information sheet found")
            if sheets =="management":
               management = book.worksheets[index]
            else:
              print("No management sheet found")
          # iterate through ground nesting sheet
          
            return self
        except:
           tb = sys.exc_info()[2]
           tbinfo = traceback.format_tb(tb)[0]
      
           # Concatenate information together concerning the error into a message string
           pymsg = "PYTHON ERRORS:\nTraceback info:\n" + tbinfo + "\nError Info:\n" + str(sys.exc_info()[1])
           print(pymsg + "\n")
      
           if arcpy.GetMessages(2) not in pymsg:
              msgs = "ArcPy ERRORS:\n" + arcpy.GetMessages(2) + "\n"
              arcpy.AddError(msgs)
              print(msgs)
      
        finally:
            end = time.perf_counter()
            print(f'reading simulation info: {end-start} seconds')
class ReadExcel2:
      def __init__(self, fullpath):
        '''
        costructs class for creating lookup info. readexcel2 takes in a full path for the excelname
        
        paramters
        ------------
        
        excelname: name with the simulation info
        
        parameters
        '''
        self.fullpath = fullpath
        if (os.path.isfile(self.fullpath)) and (self.fullpath.endswith('.xlsx')):
            pass
        else:
          logging.error("File does not exists make sure it is the specified in the working directory and no csv file is allowed")
      
      def ExtractInfo(self):
        '''
        
        '''
        try:
            start = time.perf_counter()
            self.objectid = []
            self.lonlat =[]
            self.lat  = []
              #load the excel book
            book = openpyxl.load_workbook(self.fullpath)
          # get the sheet names 
            sheet = book.sheetnames
            for index, sheets in enumerate(sheet):
             if sheets =="Lookup_info":
               lookupinfo = book.worksheets[index]
               for i in range(1, lookupinfo.max_row+1):
                    if i !=1:
                      self.objectid.append(lookupinfo.cell(row = i, column = 1).value)
                      self.lonlat.append([lookupinfo.cell(row =i, column = 2).value, lookupinfo.cell(row =i, column = 3).value])
                      self.lat.append(lookupinfo.cell(row =i, column = 3).value)  
            if "Lookup_info" not in sheet:
              print("No look up information sheet found")
            if sheets =="management":
               management = book.worksheets[index]
            else:
              print("No management sheet found")
          # iterate through ground nesting sheet
          
            return self
        except:
           tb = sys.exc_info()[2]
           tbinfo = traceback.format_tb(tb)[0]
      
           # Concatenate information together concerning the error into a message string
           pymsg = "PYTHON ERRORS:\nTraceback info:\n" + tbinfo + "\nError Info:\n" + str(sys.exc_info()[1])
           print(pymsg + "\n")
      
        finally:
            end = time.perf_counter()
            print(f'reading simulation info: {end-start} seconds')

def create_fishnet(watershedfc, height =200, SR = 4326):
    try:
        arcpy.env.scratchWorkspace = "in_memory"
        arcpy.env.overwriteOutput = True
        geodata = 'Gis_result_geodatabase.gdb'
        if not arcpy.Exists(geodata):
           arcpy.CreateFileGDB_management(os.getcwd(), geodata)
        arcpy.env.workspace = os.path.join(os.getcwd(), geodata)
        path = os.path.join(os.getcwd(), geodata)
        # change the env according to inputfc
        envcod  = arcpy.da.Describe(watershedfc)['spatialReference'].name
        arcpy.env.outputCoordinateSystem = arcpy.SpatialReference(envcod)
      # create a fishnet
        templateExtent = watershedfc
        fc= templateExtent
        labelname = 'box_fishnet'
        fcname = os.path.join(geodata, labelname)
        if arcpy.Exists(fcname):
          arcpy.management.Delete(fcname)
        desc = arcpy.Describe(watershedfc)
        in_feature_path = arcpy.CreateFishnet_management(fcname,str(desc.extent.lowerLeft),str(desc.extent.XMin) + " " + str(desc.extent.YMax + 10),
            f"{height}",f"{height}","0","0",str(desc.extent.upperRight),"NO_LABELS","#","POLYGON")
        #let's clip out the unwanted part according to the processing extent
        in_features = in_feature_path
        clip_features = watershedfc
        out_feature_class = "fcc_clipped"
        if arcpy.Exists(out_feature_class):
          arcpy.management.Delete(out_feature_class)
        out_feature_class = None
        out_feature_class = "fcc_clipped"
        #if arcpy.exists(out_feature_class):
          #arcpy.
        xy_tolerance = ""
        # Execute Clip
        arcpy.Clip_analysis(in_features, clip_features, out_feature_class, xy_tolerance)
        # convert each fishnet to point
        name = os.path.join(path, "fishnets") 
        if arcpy.Exists(name):
          arcpy.management.Delete(name)#+ str(watershedfc[-9:])
        fishnets_points = arcpy.management.FeatureToPoint(out_feature_class, name, 'INSIDE')
        sr = arcpy.SpatialReference(SR)
        listfield = []
        lf = arcpy.ListFields(fishnets_points)
        for i in lf:
            listfield.append(i.name)
        feature_array  = arcpy.da.FeatureClassToNumPyArray(name,  listfield, spatial_reference = sr)
        
        return  feature_array, fishnets_points, out_feature_class
    except  Exception as e:
       logger.exception(repr(e))
# test the code
# arcpya, bar = create_fishnet(templateExtent)

def create_point_feature_class(dat, cellSize = 0.02, field ='meanN20', srf = 4326):
  '''
  paramters:
  ---------------------
  dat: is a panda data frame
  srf = spatial reference name or code
  ------------------------
  '''
  geodata = 'Gis_result_geodatabase.gdb'
  arcpy.env.workspace = os.path.join(os.getcwd(), geodata)
  if not arcpy.Exists(geodata):
       arcpy.CreateFileGDB_management(os.getcwd(), geodata)
  point_feature_class = 'results_point_feature_class'
  data = dat #dat.groupby(['CompName']).mean()
  if arcpy.Exists(point_feature_class):
    arcpy.management.Delete(point_feature_class)
  fd  = data.convert_dtypes()
  structuredNumpy_records = data.to_records(index= False)
  fc_name = os.path.join(geodata, point_feature_class)
  srf = arcpy.SpatialReference(srf)
  # change some unassigned data type
  dt = structuredNumpy_records.dtype.descr
  for idd, elem  in zip(range(len(dt)), dt):
    if 'Soiltype' in elem:
      dt[idd] = ('Soiltype', '<U25')
    elif 'CompName' in elem:
      dt[idd] = ('CompName', '<U25')
    else:
      pass
  
  featuretobe = structuredNumpy_records.astype(dt)
  arcpy.da.NumPyArrayToFeatureClass(featuretobe, fc_name, ["longitude","Latitude"], srf)
  inFeatures= fc_name 
  outRaster = fc_name + "raster"
  if arcpy.Exists(outRaster):
    arcpy.management.Delete(outRaster)
  valField  = 'meanN20'
  priorityField = ''
  assignmentType = 'MOST_FREQUENT'
  arcpy.PointToRaster_conversion(inFeatures, valField, outRaster, 
                               assignmentType, priorityField, cellSize)
                               
  return  outRaster

def convert_simulated_to_raster(watershedfc, height =200, SR = 4326):
    geodata = 'Gis_result_geodatabase.gdb'
    gdb_data = os.path.join(os.getcwd(), geodata)
    if not arcpy.Exists(gdb_data):
       arcpy.CreateFileGDB_management(os.getcwd(), geodata)
    arcpy.env.workspace = os.path.join(os.getcwd(), geodata)
    fishnet_feature_class = 'fishnet_feature_class'
    if arcpy.Exists(fishnet_feature_class):
      arcpy.management.Delete(fishnet_feature_class)
    # change the env according to inputfc
    envcod  = arcpy.da.Describe(watershedfc)['spatialReference'].name
    arcpy.env.outputCoordinateSystem = arcpy.SpatialReference(envcod)
  # create a fishnet
    templateExtent = watershedfc
    fc= templateExtent
    fcname = fishnet_feature_class
    desc = arcpy.Describe(fc)
    in_feature_path = arcpy.CreateFishnet_management(fcname,str(desc.extent.lowerLeft),str(desc.extent.XMin) + " " + str(desc.extent.YMax + 10),f"{height}",f"{height}","0","0",str(desc.extent.upperRight),"NO_LABELS","#","POLYGON")
    #let's clip out the unwanted part according to the processing extent
    in_features = in_feature_path
    clip_features = watershedfc
    out_feature_class = 'clipped_' + fishnet_feature_class 
    #if arcpy.exists(out_feature_class):
      #arcpy.
    xy_tolerance = ""
    if arcpy.Exists(out_feature_class):
      arcpy.management.Delete(out_feature_class)
    # Execute Clip
    arcpy.Clip_analysis(in_features, clip_features, out_feature_class, xy_tolerance)
    InZones = out_feature_class
    InZoneField = "OID"
    InValueRaster = "rras.tif"
    
    # Check out ArcGIS Spatial Analyst extension license
    arcpy.CheckOutExtension("Spatial")
    
    # Process: Calculate the mean slope of each parcel area.
    out = arcpy.sa.ZonalStatistics(InZones, InZoneField, InValueRaster, "MEAN", 
                                   "DATA")
    # convert each fishnet t
# this create a rasterlayer
def create_results_raster_layer(dat, inraster, cellSize = 0.02, field ='meanMaizeAGB',  assignmentType = 'MOST_FREQUENT'):
  '''
  creates raster layer from point simulated data
  
  paramters
  ------------------
  inraster: same as the soil raster or similar to the simulated extent
  dat: pandas data frame
  field: field name to be used in create contnous raster data
  cellsise: size to calcualte the statics e.g mean
  -------------------------------
  returns a string path of the out rastername and poit feature classpath
  '''
  srf = 4326
  arcpy.env.snapRaster = inraster
  envcod  = arcpy.da.Describe(inraster)['spatialReference'].name
  arcpy.env.outputCoordinateSystem = arcpy.SpatialReference(envcod)
  '''
  paramters:
  ---------------------
  dat: is a panda data frame
  srf = spatial reference name or code
  ------------------------
  '''
  geodata = 'Gis_result_geodatabase.gdb'
  arcpy.env.workspace = os.path.join(os.getcwd(), geodata)
  if not arcpy.Exists(geodata):
       arcpy.CreateFileGDB_management(os.getcwd(), geodata)
  point_feature_class = 'results_point_feature_class'
  ptfc = os.path.join(os.path.join(os.getcwd(), geodata), point_feature_class)
  data = dat #dat.groupby(['CompName']).mean()
  if arcpy.Exists(ptfc):
    arcpy.management.Delete(ptfc)
  fd  = data.convert_dtypes()
  structuredNumpy_records = data.to_records(index= False)
  fc_name = os.path.join(geodata, point_feature_class)
  srf = arcpy.SpatialReference(srf)
  # change some unassigned data type
  dt = structuredNumpy_records.dtype.descr
  for idd, elem  in zip(range(len(dt)), dt):
    if 'Soiltype' in elem:
      dt[idd] = ('Soiltype', '<U25')
    elif 'CompName' in elem:
      dt[idd] = ('CompName', '<U25')
    else:
      pass
  featuretobe = structuredNumpy_records.astype(dt)
  arcpy.da.NumPyArrayToFeatureClass(featuretobe, fc_name, ["longitude","Latitude"], envcod)
  inFeatures= fc_name
  outRaster = None
  outRaster = str(field) + "ras"
  if arcpy.Exists(outRaster):
    arcpy.management.Delete(outRaster)
  valField  = field
  priorityField = ''
  cellSize = cellSize
  arcpy.PointToRaster_conversion(inFeatures, valField, outRaster, 
                               assignmentType, priorityField, cellSize)
                               
  return  outRaster, fc_name

def extract_param_table(string):
  data_dict = {}
  str_list   = string.split(";")
  i = None
  for i in str_list:
    if 'Maize, Wheat, Soybean' in i:
      float_num = float(i[-3:])
      data_dict["Maize_wheat_rye"] = ["Maize, Wheat, Soybean", float_num] 
    elif 'Maize' == i.split(" ")[0]:
      float_num = float(i[-3:])
      data_dict["Maize_only"] = ["Maize", float_num] 
    elif 'Maize, Wheat' in i:
      float_num = float(i[-3:])
      data_dict["maize_rye_only"] = ["Maize, Wheat", float_num]
    elif 'Maize, Soybean' == i.split("'")[1]:
      float_num = float(i[-3:])
      data_dict["maize_soybean_only"] = ["Maize, Soybean", float_num]
  return data_dict
# deleting simulated files
def delete_simulation_files(path):
  weather_files_path = opj(path, 'weatherdata')
  weather_files = glob.glob1(weather_files_path, 'weather_Daymet*.met')
  patterns = ['*lat*lon*.csv', 'np*.txt', 'edit*.apsimx', 'edit*.db-shm', 'edit*.db-wal', 'edit*.db', 'fish*.cpg', 'edit*.bak', 'coh*.apsimx', 'coh*.bak', 'fish*.shp', 'fish*.shx']
  localfiles = [glob.glob1(path, pat) for pat in patterns]
  files_to_delete = []
  for file_group in localfiles:
    absolute_paths = [opj(path, file) for file in file_group]
    files_to_delete.extend(absolute_paths)
  for i in files_to_delete:
    try:
      os.remove(i)
    except PermissionError as e:
      print(repr(e))
def delete_weather_files(path):
  weather_files_path = opj(path, 'weatherdata')
  weather_files = glob.glob1(weather_files_path, 'weath*.met')
  files_2_delete = [opj(weather_files_path, fi) for fi in weather_files]
  for i in files_2_delete:
    try:
      os.remove(i)
    except PermissionError as e:
      print(repr(e))   
# saving simulated files
def save_simulation_results(result_list, file_name = "resultsexternal.csv"):
  data_df = makedf(result_list)
  print(data_df.head())
  print(data_df.shape)
  base_sim_path  = os.path.join(os.getcwd(), "SimulationResults")
  if not os.path.exists(base_sim_path):
          os.mkdir(base_sim_path)
  results_csv = os.path.join(base_sim_path, file_name)
  if os.path.isfile(results_csv):
      os.remove(results_csv)
  results_csv = os.path.join(base_sim_path, file_name)
  data_df.to_csv(results_csv)
# sampling
def sample_list_use_dict(lst, paramatable):
    """
    takes in a list or a tuple and sample elements according to the specified probabilities
    we use random.sample which samples without replacement and then for each sample
    we remove the sampled elements from the list
    
    --- parameters
    paramatable is adictionary of extracted from the user inputs
    lst is a list of  indices for the numpy arrays
    """
    dp = paramatable
    list_len = len(lst)
    prob = []
    for i in dp.keys():
      prob.append(dp[i][1])
    
    crops = []
    elem  = None
    for elem in dp.keys():
      crops.append(dp[elem][0])
    crop_dist = {}
    for i, m, crop in zip(range(len(crops)), prob, crops):
       crop_dist[crop] = int(m*list_len)
    #create a dictionary and insert a dummy key
    data = {"crop":(0)}
    for i, crop in zip(crop_dist.items(), crops):
      if data["crop"]:
        for i in data['crop']:
          if i in lst:
            lst.remove(i)
      data[crop] = random.sample(lst, crop_dist[crop])
    # remove the dummy
    data.pop('crop')
    return data
  
def printfieldname(fc):
      listfield = []
      lf = arcpy.ListFields(fc)
      for i in lf:
        listfield.append(i.name)
      print(listfield)
#mapping the 
class result_Gis_Management:
  def __init__(self, in_raster, horizontable, data_frame = '', rasterfield = "MUKEY", cellSize = 0.02, 
                  field ='meanMaizeAGB',  assignmentType = 'MOST_FREQUENT'):
    self.raster = in_raster
    self.field = rasterfield
    self.Horizontable = horizontable
    self.data_frame = data_frame
    self.cellSize = cellSize
    self.field2 = field
    self.assignmentType = assignmentType
  def Organise_output(self, cellsize =30):
    try:
      geodata = 'Gis_result_geodatabase.gdb'
      if not arcpy.Exists(geodata):
       arcpy.CreateFileGDB_management(os.getcwd(), geodata)
      arcpy.env.workspace  = geodata
      shapename = r'in_memory/shapelayer'
      if arcpy.Exists(shapename):
        arcpy.management.Delete(shapename)
      shapename = r'in_memory/shapelayer'
      #arcpy.AddMessage('converting raster to polygon')
      self.rasterpolygon =  arcpy.conversion.RasterToPolygon(self.raster , shapename, simplify ="NO_SIMPLIFY", raster_field =self.field)
      printfieldname(self.rasterpolygon)
      #Process results to feature class
      in_raster_path = None
      in_raster_path = self.raster
      raster_layer_pathx, pf_class = create_results_raster_layer(self.data_frame, in_raster_path, self.cellSize, 
                            self.field2,  self.assignmentType)
      
      name = r'in_memory/feature_to_point'
      if arcpy.Exists(name):
        arcpy.management.Delete(name)
      # redefine it
      name = r'in_memory/feature_to_point'
      print('converting polygon to points')
      #arcpy.AddMessage('converting Polygon to points')
      # convert to points
      self.rasterpoints = arcpy.management.FeatureToPoint(self.rasterpolygon , name, 'INSIDE')
      inFeatures =self.rasterpolygon # self.rasterpoints
      joinTable = self.Horizontable
      joinField = "OBJECTID"
      expression = "CompKind = 'Series'"
      outFeature = "spatialsoilhorizon"
      print("joining tables")
      self.joinedtable = arcpy.management.AddJoin(inFeatures, joinField, pf_class, joinField)
      ap = os.path.join
      arcpy.MakeFeatureLayer_management(self.joinedtable, "fc1_lyr12")                                        
      sr = arcpy.SpatialReference(4326)
      return self
    except  Exception as e:
      logger.exception(repr(e))
# import os
# data = os.path.join(r'C:\Users\rmagala\Box\Simulation_Application\Papsimx\SimulationResults', 'Continous_maize_acpf070801050101.csv')
# import pandas as pd
# df = pd.read_csv(data)
# inraster = r'C:\ACPd\Base_files\acpf_huc070801050101\acpf070801050101.gdb\gSSURGO'
# table  = r'C:\ACPd\Base_files\acpf_huc070801050101\acpf070801050101.gdb\SurfHrz070801050101'
# jtable = result_Gis_Management(inraster, table, df).Organise_output()

#ppp =SoilRasterManagement(inraster, table).Organise_soils()
